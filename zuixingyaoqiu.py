import pandas as pd
import numpy as np
import faiss
from langchain_community.embeddings import DashScopeEmbeddings
from langchain_community.chat_models import ChatTongyi
from langchain_community.vectorstores import FAISS
from langchain_openai import ChatOpenAI
import re
from typing import List, Dict, Tuple, Optional
import os
import datetime
# 使用本地部署的向量化模型
from langchain_ollama import OllamaEmbeddings
#设置OpenAI环境变量
os.environ["OPENAI_API_KEY"] = "a2nR7TLLqbBEtfZy8xaMyf47DJ_V04ZhlOUWA68R7vs"
os.environ["OPENAI_BASE_URL"] = "http://140.210.143.2:57000/vllm/v1"

class EnterpriseMatchingSystem:
    def __init__(self, template_path: str, dashscope_api_key: str = None):
        """
        初始化企业匹配系统
        """
        self.template_path = template_path
        self.dashscope_api_key = dashscope_api_key
        self.embedding_model = None
        self.llm = None
        self.faiss_index = None
        self.company_names = []
        
        # 初始化模型
        self._initialize_models()
        
        # 加载结果模板
        self.template_df = self._load_template()
        
        # 临时存储表：包含企业名称、信用代码、在模板中的行号
        self.temp_table = self._initialize_temp_table()

 
    def _initialize_models(self):
        """初始化嵌入模型和LLM"""
        try:
            
            
            # 初始化嵌入模型 - 使用DashScope模型
            self.embedding_model = DashScopeEmbeddings(
                model="text-embedding-v4",
                dashscope_api_key=self.dashscope_api_key
            )
            """

            # 初始化嵌入模型 - 使用本地部署的Ollama模型
            self.embedding_model = OllamaEmbeddings(
                model="quentinz/bge-large-zh-v1.5:q4_0",  # 就是你 pull 的那个名字
                base_url="http://192.168.2.172:11434"
      # Windows Ollama 使用这个地址
            )
           """
            # 初始化大模型 - 使用OpenAI兼容接口
            openai_key = os.getenv("OPENAI_API_KEY", "")
            openai_base = os.getenv("OPENAI_BASE_URL", "http://140.210.143.2:57000/vllm/v1")
            
            if openai_key and openai_base:
                self.llm = ChatOpenAI(
                    api_key=openai_key, 
                    base_url=openai_base, 
                    model="Qwen2.5-72B-Instruct"
                )
                print("OpenAI模型初始化成功")
            else:
                print("OpenAI API密钥或基础URL未设置，使用通义千问作为备选")
                self.llm = ChatTongyi(
                    model_name="qwen-max",
                    dashscope_api_key=self.dashscope_api_key
                )
                
        except Exception as e:
            print(f"模型初始化失败: {e}")
            print("请检查API密钥是否正确，以及是否已安装所需包")
            self.embedding_model = None
            self.llm = None
    
    def _load_template(self) -> pd.DataFrame:
        """加载结果模板"""
        try:
            # 读取模板，不指定header以便正确处理多行表头
            df = pd.read_excel(self.template_path, sheet_name='Sheet1', header=None)
            print(f"模板加载成功，共{len(df)}行数据，{len(df.columns)}列")
            
            # 删除前两行表头，保留数据行
            data_df = df.iloc[2:].reset_index(drop=True)
            # 设置列名
            data_df.columns = list(range(len(df.columns)))
            target_idx = 16
            if target_idx <= len(data_df.columns):
                data_df.insert(target_idx, '匹配结果', '')
                self.status_col_index = target_idx
            else:
                data_df['匹配结果'] = ''
                self.status_col_index = len(data_df.columns) - 1
            return data_df
        except Exception as e:
            print(f"模板加载失败: {e}")
            return pd.DataFrame()
    
    def _initialize_temp_table(self) -> List[Dict]:
        """初始化临时表，从模板的在线监测数据开始"""
        temp_table = []
        
        for idx, row in self.template_df.iterrows():
            # 根据实际列位置，企业名称在第3列(索引3)，统一社会信用代码在第4列(索引4)
            # 修改：只要求企业名称非空，不要求信用代码非空
            if pd.notna(row.iloc[3]) and str(row.iloc[3]).strip() != '':
                temp_table.append({
                    'name': str(row.iloc[3]).strip(),
                    'credit_code': str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else '',
                    'template_row': idx,
                    'source': '在线监测'
                })
        
        print(f"临时表初始化完成，共{len(temp_table)}家企业")
        return temp_table
    
    def clean_company_name(self, name: str) -> str:
        """清洗企业名称"""
        if pd.isna(name):
            return ""
        name = str(name).strip()
        name = name.replace('（', '(').replace('）', ')')
        name = re.sub(r'\s+', ' ', name).strip()
        return name
    
    def exact_match(self, target_companies: List[Dict], source_companies: List[Dict]) -> Tuple[List[Tuple], List[Dict]]:
        """
        精确匹配：信用代码匹配 + 名称匹配
        返回：(匹配对列表, 未匹配的企业列表)
        """
        matched_pairs = []
        unmatched_source = source_companies.copy()
        
        # 第一阶段：信用代码精确匹配
        temp_matched = []
        for temp_company in target_companies:
            for source_idx, source_company in enumerate(unmatched_source):
                if (temp_company['credit_code'] and source_company['credit_code'] and 
                    temp_company['credit_code'] == source_company['credit_code']):
                    matched_pairs.append((temp_company, source_company))
                    temp_matched.append(source_idx)
                    break
        
        # 移除已匹配的
        unmatched_source = [company for idx, company in enumerate(unmatched_source) 
                           if idx not in temp_matched]
        
        # 第二阶段：名称精确匹配或包含匹配
        temp_matched = []
        for temp_company in target_companies:
            # 跳过已经通过信用代码匹配的企业
            if temp_company in [pair[0] for pair in matched_pairs]:
                continue
                
            clean_temp_name = self.clean_company_name(temp_company['name'])
            for source_idx, source_company in enumerate(unmatched_source):
                clean_source_name = self.clean_company_name(source_company['name'])
                
                if clean_temp_name == clean_source_name:
                    matched_pairs.append((temp_company, source_company))
                    temp_matched.append(source_idx)
                    break
        
        # 移除已匹配的（按索引从大到小排序后移除，避免索引错位）
        temp_matched.sort(reverse=True)
        for idx in temp_matched:
            if idx < len(unmatched_source):
                unmatched_source.pop(idx)
        
        return matched_pairs, unmatched_source
    
    def _log_vector_matches(self, vector_matched: List[Dict], matched_pairs: List[Tuple], source_type: str):
        """记录向量匹配结果到文件"""
        try:
            # 创建日志目录
            log_dir = os.path.join(os.path.dirname(self.template_path), "匹配日志")
            os.makedirs(log_dir, exist_ok=True)
            
            # 生成日志文件名
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            log_file = os.path.join(log_dir, f"向量匹配结果_{timestamp}.txt")
            
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write(f"向量匹配结果记录 - {timestamp}\n")
                f.write(f"匹配来源: {source_type}\n")
                f.write(f"匹配成功数量: {len(vector_matched)}家\n")
                f.write("=" * 50 + "\n\n")
                
                for i, matched_company in enumerate(vector_matched, 1):
                    # 找到对应的模板企业
                    template_company = None
                    for pair in matched_pairs:
                        if pair[1] == matched_company:
                            template_company = pair[0]
                            break
                    
                    f.write(f"第{i}组 - 向量匹配成功:\n")
                    if template_company:
                        f.write(f"  模板企业: {template_company['name']}\n")
                        f.write(f"  模板信用代码: {template_company.get('credit_code', '')}\n")
                    f.write(f"  匹配企业: {matched_company['name']}\n")
                    f.write(f"  匹配信用代码: {matched_company.get('credit_code', '')}\n")
                    f.write(f"  来源文件: {matched_company.get('source_file', '未知')}\n")
                    f.write("-" * 30 + "\n\n")
            
            print(f"向量匹配结果已保存到: {log_file}")
            
        except Exception as e:
            print(f"记录向量匹配结果失败: {e}")
    
    def build_vector_index(self, company_names: List[str]):
        """构建企业名称的向量索引"""
        if not self.embedding_model:
            print("嵌入模型未初始化，无法进行向量匹配")
            return None
        
        try:
            # 使用FAISS创建向量数据库
            if company_names:
                start_time = datetime.datetime.now()
                self.vector_store = FAISS.from_texts(
                    company_names, 
                    self.embedding_model
                )
                self.company_names = company_names
                elapsed = (datetime.datetime.now() - start_time).total_seconds()
                print(f"向量索引构建完成，共{len(company_names)}个企业名称, "
                      f"耗时: {elapsed:.2f}秒, "
                      f"平均每个: {elapsed/len(company_names):.4f}秒")
                return self.vector_store
            else:
                print("企业名称列表为空，无法构建向量索引")
                return None
            
        except Exception as e:
            print(f"向量索引构建失败: {e}")
            return None
    
    def vector_match(self, target_name: str, source_companies: List[Dict], top_k: int = 3) -> List[Tuple[Dict, float]]:
        """向量相似度匹配"""
        if not hasattr(self, 'vector_store') or not self.vector_store:
            return [] 
        
        try:
            # 搜索最相似的top_k个结果
            docs = self.vector_store.similarity_search_with_score(target_name, k=top_k)
            
            results = []
            for doc, score in docs:
                # 找到对应的源企业
                for source_company in source_companies:
                    clean_source_name = self.clean_company_name(source_company['name'])
                    if clean_source_name == doc.page_content:
                        # 将距离转换为相似度（距离越小相似度越高）
                        similarity = 1.0 / (1.0 + score)
                        results.append((source_company, similarity))
                        break
            
            return results
            
        except Exception as e:
            print(f"向量匹配失败: {e}")
            return []
    
    def batch_llm_verify_match(self, verification_batch: List[Tuple[Dict, List[Tuple[Dict, float]]]]) -> List[Optional[Dict]]:
        """批量使用大模型验证匹配结果：选出3个最相似的候选，让大模型判断哪个与目标企业是同一家公司"""
        if not self.llm:
            print("大模型未初始化，使用相似度最高的结果")
            return [candidates[0][0] if candidates else None for _, candidates in verification_batch]
        
        # 记录候选者信息到日志文件
        self._log_candidates(verification_batch)
        
        try:
            # 构建批量提示词 - 让大模型从3个候选中选择哪个与目标企业是同一家公司
            batch_prompt = "你是一个熟悉四川省行政区划和企业命名规范的专家，任务是从每组候选企业中准确识别出与目标企业为同一家公司的选项。\n"
            batch_prompt += "请从每组候选企业中选出与目标企业是同一家公司的那个，不同文件中同一家企业可能有不同的名称表示：\n"
            batch_prompt += "**重要规则（必须严格遵守）**: \n"
            batch_prompt += "1. 地域规则：地区名称不同绝对不是同一家公司，必须严格检查市县区镇乡等地名，不同行政区划绝对不匹配\n"
            batch_prompt += "2. 核心业务规则：核心业务名称必须相同或高度相似才可能是同一家公司\n"
            batch_prompt += "3. 地域优先级：地域匹配的优先级高于名称相似度，即使名称相似度很高，地域不同也绝对不是同一家公司\n"
            batch_prompt += "4. 四川地域知识：运用上述四川省地区知识进行精确判断，注意四川特有的市县区名称\n"
            batch_prompt += "5. 严格检查：必须逐个字符检查地域名称，确保完全一致，不同市县区绝对不能匹配\n\n"
            
            for i, (target_company, candidate_companies) in enumerate(verification_batch):
                batch_prompt += f"第{i+1}组：\n"
                batch_prompt += f"目标企业：{target_company['name']}\n"
                
                batch_prompt += "候选企业（已按相似度排序）：\n"
                for j, (candidate, score) in enumerate(candidate_companies):
                    batch_prompt += f"  {j+1}. {candidate['name']}\n"
                
                batch_prompt += f"请返回数字（1-{len(candidate_companies)}）表示哪个候选企业与目标企业是同一家公司，如果没有匹配的返回0。\n\n"
            
            # 添加响应格式说明
            batch_prompt += "请按顺序返回每组的结果，用逗号分隔，例如：1,0,3,2"
            
            response = self.llm.invoke(batch_prompt)
            response_text = response.content if hasattr(response, 'content') else str(response)
            
            # 解析批量结果
            results = []
            match_results = re.findall(r'\d+', response_text)
            
            for i, (target_company, candidate_companies) in enumerate(verification_batch):
                if i < len(match_results):
                    match_idx = int(match_results[i]) - 1
                    if 0 <= match_idx < len(candidate_companies):
                        results.append(candidate_companies[match_idx][0])
                    else:
                        results.append(None)
                else:
                    results.append(None)
            
            return results
            
        except Exception as e:
            print(f"批量大模型验证失败: {e}")
            # 失败时返回相似度最高的结果
            return [candidates[0][0] if candidates else None for _, candidates in verification_batch]
    
    def llm_verify_match(self, target_company: Dict, candidate_companies: List[Tuple[Dict, float]]) -> Optional[Dict]:
        """兼容原有的单个验证方法"""
        results = self.batch_llm_verify_match([(target_company, candidate_companies)])
        return results[0] if results else None
    
    def _log_llm_selection(self, target_company: Dict, candidate_companies: List[Tuple[Dict, float]], choice: Optional[Dict]):
        try:
            log_dir = "/home/wjw/diting/ZhengHe/xingxuqiu/匹配日志"
            os.makedirs(log_dir, exist_ok=True)
            log_file = os.path.join(log_dir, "大模型选择结果日志.txt")
            if not os.path.exists(log_file):
                with open(log_file, 'w', encoding='utf-8') as f:
                    f.write("=" * 80 + "\n")
                    f.write(f"大模型选择结果日志 - 程序开始时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("=" * 80 + "\n\n")
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"目标企业: {target_company['name']}\n")
                f.write("候选企业（已按相似度排序）：\n")
                for idx, (cand, _) in enumerate(candidate_companies, 1):
                    f.write(f"  {idx}. {cand['name']}\n")
                f.write(f"选择结果: {choice['name'] if choice else '无匹配(0)'}\n")
                f.write("-" * 60 + "\n\n")
        except Exception as e:
            print(f"记录大模型选择结果失败: {e}")
    
    def _log_candidates(self, verification_batch: List[Tuple[Dict, List[Tuple[Dict, float]]]]):
        """记录向量匹配候选者信息到日志文件（已按相似度排序的前3个候选）"""
        try:
            # 创建日志目录
            log_dir = "/home/wjw/diting/ZhengHe/xingxuqiu/匹配日志"
            os.makedirs(log_dir, exist_ok=True)
            
            # 使用固定的日志文件名，每次运行覆盖之前的日志
            log_file = os.path.join(log_dir, "候选者匹配日志.txt")
            
            # 如果是第一次写入，先写入文件头
            if not os.path.exists(log_file):
                with open(log_file, 'w', encoding='utf-8') as f:
                    f.write(f"=" * 80 + "\n")
                    f.write(f"候选者匹配日志 - 程序开始时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"=" * 80 + "\n\n")
            
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"批次记录时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"本批匹配组数: {len(verification_batch)}\n")
                f.write(f"{'=' * 60}\n\n")
                
                for i, (target_company, candidate_companies) in enumerate(verification_batch):
                    f.write(f"第{i+1}组 - 目标企业:\n")
                    f.write(f"  名称: {target_company['name']}\n")
                    f.write(f"  候选企业数量: {len(candidate_companies)}\n")
                    f.write(f"  {'-' * 60}\n")
                    
                    for j, (candidate, _) in enumerate(candidate_companies):
                        f.write(f"  候选{j+1}:\n")
                        f.write(f"    名称: {candidate['name']}\n")
                        f.write(f"\n")
                    
                    f.write(f"\n")
                
                f.write(f"{'=' * 80}\n\n")
            
        except Exception as e:
            print(f"记录候选者日志失败: {e}")
    
    def match_companies(self, source_file: str, source_type: str) -> List[Dict]:
        """
        匹配企业信息
        source_type: '用电监控', '排污许可', '减排清单', '源清单'
        """
        print(f"\n开始匹配{source_type}数据...")
        
        # 加载源数据
        source_df = self._load_source_data(source_file, source_type)
        if source_df.empty:
            return self.temp_table
        
        # 提取源企业信息
        source_companies = self._extract_source_companies(source_df, source_type)
        print(f"源数据包含{len(source_companies)}家企业")
        
        # 精确匹配
        matched_pairs, unmatched_source = self.exact_match(self.temp_table, source_companies)
        print(f"精确匹配成功: {len(matched_pairs)}家")
        
        # 向量匹配（针对未匹配的企业）
        if unmatched_source and self.embedding_model:
            # 构建源企业名称列表用于向量匹配
            source_names = [self.clean_company_name(company['name']) for company in unmatched_source]
            self.build_vector_index(source_names)
            
            vector_matched = []
            verification_batch = []  # 用于批量验证
            
            # 向量匹配耗时统计
            vector_start_time = datetime.datetime.now()
            vector_match_count = 0
            
            # 计算需要匹配的总企业数
            temp_companies_to_match = []
            for temp_company in self.temp_table:
                if not any(temp_company == pair[0] for pair in matched_pairs):
                    temp_companies_to_match.append(temp_company)
            
            total_to_match = len(temp_companies_to_match)
            
            if total_to_match > 0:
                print(f"开始向量匹配，共{total_to_match}家企业需要处理...")
            
            # 收集需要验证的匹配对
            for idx, temp_company in enumerate(temp_companies_to_match):
                # 向量匹配
                candidates = self.vector_match(
                    self.clean_company_name(temp_company['name']), 
                    unmatched_source
                )
                
                if candidates:
                    verification_batch.append((temp_company, candidates))
                    vector_match_count += 1
                
                # 实时进度显示（每10个企业或最后1个）
                current_progress = idx + 1
                if current_progress % 10 == 0 or current_progress == total_to_match:
                    elapsed = (datetime.datetime.now() - vector_start_time).total_seconds()
                    progress_percent = (current_progress / total_to_match) * 100
                    
                    # 计算预计剩余时间
                    if current_progress > 0:
                        time_per_item = elapsed / current_progress
                        remaining_time = time_per_item * (total_to_match - current_progress)
                        eta_str = f", 预计剩余: {remaining_time:.1f}秒"
                    else:
                        eta_str = ""
                    
                    # 进度条显示
                    bar_length = 20
                    filled_length = int(bar_length * current_progress // total_to_match)
                    bar = "█" * filled_length + "░" * (bar_length - filled_length)
                    
                    print(f"向量匹配进度: [{bar}] {progress_percent:.1f}% "
                          f"({current_progress}/{total_to_match}), "
                          f"已耗时: {elapsed:.1f}秒{eta_str}")
            
            if total_to_match > 0:
                vector_elapsed = (datetime.datetime.now() - vector_start_time).total_seconds()
                print(f"向量匹配完成: 共{vector_match_count}个匹配对, "
                      f"总耗时: {vector_elapsed:.2f}秒, "
                      f"平均每个: {vector_elapsed/max(1, vector_match_count):.4f}秒")
            
            # 批量处理验证（每次处理40个）
            batch_size = 40
            llm_start_time = datetime.datetime.now()
            total_batches = (len(verification_batch) + batch_size - 1) // batch_size
            
            if total_batches > 0:
                print(f"开始大模型验证，共{total_batches}个批次需要处理...")
            
            for batch_idx, i in enumerate(range(0, len(verification_batch), batch_size)):
                batch = verification_batch[i:i+batch_size]
                batch_start_time = datetime.datetime.now()
                
                batch_results = self.batch_llm_verify_match(batch)
                
                batch_elapsed = (datetime.datetime.now() - batch_start_time).total_seconds()
                
                for j, (temp_company, candidates) in enumerate(batch):
                    choice = batch_results[j] if j < len(batch_results) else None
                    if choice:
                        matched_company = choice
                        matched_pairs.append((temp_company, matched_company))
                        unmatched_source = [company for company in unmatched_source 
                                         if company != matched_company]
                        vector_matched.append(matched_company)
                    self._log_llm_selection(temp_company, candidates, choice)
                
                # 实时进度显示
                current_batch = batch_idx + 1
                elapsed = (datetime.datetime.now() - llm_start_time).total_seconds()
                progress_percent = (current_batch / total_batches) * 100
                
                # 计算预计剩余时间
                if current_batch > 0:
                    time_per_batch = elapsed / current_batch
                    remaining_time = time_per_batch * (total_batches - current_batch)
                    eta_str = f", 预计剩余: {remaining_time:.1f}秒"
                else:
                    eta_str = ""
                
                # 进度条显示
                bar_length = 20
                filled_length = int(bar_length * current_batch // total_batches)
                bar = "█" * filled_length + "░" * (bar_length - filled_length)
                
                print(f"大模型验证进度: [{bar}] {progress_percent:.1f}% "
                      f"({current_batch}/{total_batches}), "
                      f"本批耗时: {batch_elapsed:.1f}秒, "
                      f"总耗时: {elapsed:.1f}秒{eta_str}")
            
            if total_batches > 0:
                llm_elapsed = (datetime.datetime.now() - llm_start_time).total_seconds()
                print(f"大模型验证完成: 共{len(verification_batch)}组, "
                      f"总耗时: {llm_elapsed:.2f}秒, "
                      f"平均每组: {llm_elapsed/max(1, len(verification_batch)):.2f}秒")
            
            # 记录向量匹配结果到文件
            if vector_matched:
                print(f"向量匹配成功: {len(vector_matched)}家")
                self._log_vector_matches(vector_matched, matched_pairs, source_type)
            else:
                print(f"向量匹配成功: 0家")
        
        # 更新模板数据
        self._update_template(matched_pairs, source_type)
        
        # 添加未匹配的源企业到临时表
        new_rows_start = len(self.template_df)
        for i, company in enumerate(unmatched_source):
            new_row_idx = new_rows_start + i
            self.temp_table.append({
                'name': company['name'],
                'credit_code': company['credit_code'],
                'template_row': new_row_idx,
                'source': source_type
            })
        
        # 在模板中添加新行
        self._add_new_rows_to_template(unmatched_source, source_type)
        
        print(f"{source_type}匹配完成: 总匹配{len(matched_pairs)}家，新增{len(unmatched_source)}家")
        return self.temp_table
    
    def _load_source_data(self, file_path: str, source_type: str) -> pd.DataFrame:
        """加载源数据文件"""
        try:
            # 源清单可能没有统一社会信用代码，只读取企业名称
            if source_type == '源清单':
                df = pd.read_excel(file_path)  # 不指定sheet_name，让pandas自动选择第一个工作表
                # 假设源清单只有企业名称，没有统一社会信用代码
                if '企业名称' not in df.columns:
                    # 尝试找到包含企业名称的列
                    for col in df.columns:
                        if '企业' in col or '名称' in col:
                            df = df.rename(columns={col: '企业名称'})
                            break
            else:
                df = pd.read_excel(file_path)  # 不指定sheet_name，让pandas自动选择第一个工作表
            
            # 根据文件类型调整列名映射
            if source_type == '排污许可':
                if 'entername' in df.columns:
                    df = df.rename(columns={
                        'entername': '企业名称'
                    })

            return df
        except Exception as e:
            print(f"加载{source_type}数据失败: {e}")
            return pd.DataFrame()
    
    def _extract_source_companies(self, source_df: pd.DataFrame, source_type: str) -> List[Dict]:
        """从源数据中提取企业信息"""
        companies = []
        
        # 源清单只有企业名称，没有统一社会信用代码
        if source_type == '源清单':
            name_col = '企业名称'
            for idx, row in source_df.iterrows():
                val = row.get(name_col, '')
                name = next((v for v in (val if isinstance(val, pd.Series) else [val]) if pd.notna(v) and str(v).strip() != ''), '')
                if name:
                    company_info = {
                        'name': str(name).strip(),
                        'credit_code': '',  # 源清单没有信用代码
                        'source_file': source_type,
                        'original_index': idx
                    }
                    companies.append(company_info)
        else:
            # 其他文档有企业名称和统一社会信用代码
            name_col = '企业名称'
            code_col = '统一社会信用代码'
            
            for idx, row in source_df.iterrows():
                val_name = row.get(name_col, '')
                name = next((v for v in (val_name if isinstance(val_name, pd.Series) else [val_name]) if pd.notna(v) and str(v).strip() != ''), '')
                val_code = row.get(code_col, '')
                credit_code = next((v for v in (val_code if isinstance(val_code, pd.Series) else [val_code]) if pd.notna(v) and str(v).strip() != ''), '')
                if name:
                    company_info = {
                        'name': str(name).strip(),
                        'credit_code': str(credit_code).strip(),
                        'source_file': source_type,
                        'original_index': idx
                    }
                    if source_type == '排污许可':
                        val_per = row.get('排污许可证编号', '')
                        permit_num = next((v for v in (val_per if isinstance(val_per, pd.Series) else [val_per]) if pd.notna(v) and str(v).strip() != ''), '')
                        company_info['permit_number'] = permit_num
                    companies.append(company_info)
        
        return companies
    
    def _update_template(self, matched_pairs: List[Tuple], source_type: str):
        """更新模板中的匹配结果"""
        # 根据实际列位置调整列映射
        # H列(索引7): 用电监控-企业名称
        # I列(索引8): 用电监控-统一社会信用代码
        column_mapping = {
            '用电监控': {'name_col': 7, 'code_col': 8},  # H列和I列
            '排污许可': {'name_col': 9, 'code_col': 10, 'permit_col': 11},  # J、K、L列
            '减排清单': {'name_col': 12, 'code_col': 13},  # M、N列
            '源清单': {'name_col': 14}   # O列（源清单只有企业名称）
        }
        
        if source_type not in column_mapping:
            print(f"未知的源类型: {source_type}")
            return
        
        cols = column_mapping[source_type]
        
        for temp_company, source_company in matched_pairs:
            row_idx = temp_company['template_row']
            
            # 更新企业名称
            if 'name_col' in cols:
                self.template_df.iloc[row_idx, cols['name_col']] = source_company['name']
            
            # 更新统一社会信用代码（源清单除外）
            if 'code_col' in cols and source_type != '源清单':
                self.template_df.iloc[row_idx, cols['code_col']] = source_company['credit_code']
            
            # 对于排污许可，更新许可证编号
            if source_type == '排污许可' and 'permit_col' in cols:
                permit_num = source_company.get('permit_number', '')
                self.template_df.iloc[row_idx, cols['permit_col']] = permit_num
            
            if hasattr(self, 'status_col_index'):
                if (str(temp_company['name']).strip() == str(source_company['name']).strip() and 
                    str(temp_company['credit_code']).strip() == str(source_company['credit_code']).strip()):
                    self.template_df.iloc[row_idx, self.status_col_index] = "完全匹配"
                else:
                    self.template_df.iloc[row_idx, self.status_col_index] = "不完全匹配"
            # 设置匹配结果
            if hasattr(self, 'status_col_index') and self.status_col_index is not None:
                if (str(temp_company['name']).strip() == str(source_company['name']).strip() and 
                    str(temp_company['credit_code']).strip() == str(source_company['credit_code']).strip()):
                    self.template_df.iloc[row_idx, self.status_col_index] = "完全匹配"
                else:
                    self.template_df.iloc[row_idx, self.status_col_index] = "不完全匹配"
            
            if hasattr(self, 'status_col_index'):
                if (str(temp_company['name']).strip() == str(source_company['name']).strip() and 
                    str(temp_company['credit_code']).strip() == str(source_company['credit_code']).strip()):
                    self.template_df.iloc[row_idx, self.status_col_index] = "完全匹配"
                else:
                    self.template_df.iloc[row_idx, self.status_col_index] = "不完全匹配"
            
            if hasattr(self, 'status_col_index'):
                if (str(temp_company['name']).strip() == str(source_company['name']).strip() and 
                    str(temp_company['credit_code']).strip() == str(source_company['credit_code']).strip()):
                    self.template_df.iloc[row_idx, self.status_col_index] = "完全匹配"
                else:
                    self.template_df.iloc[row_idx, self.status_col_index] = "不完全匹配"
            
            if hasattr(self, 'status_col_index'):
                if (str(temp_company['name']).strip() == str(source_company['name']).strip() and 
                    str(temp_company['credit_code']).strip() == str(source_company['credit_code']).strip()):
                    self.template_df.iloc[row_idx, self.status_col_index] = "完全匹配"
                else:
                    self.template_df.iloc[row_idx, self.status_col_index] = "不完全匹配"
            # 设置匹配结果
            if hasattr(self, 'status_col_index') and self.status_col_index is not None:
                if (str(temp_company['name']).strip() == str(source_company['name']).strip() and 
                    str(temp_company['credit_code']).strip() == str(source_company['credit_code']).strip()):
                    self.template_df.iloc[row_idx, self.status_col_index] = "完全匹配"
                else:
                    self.template_df.iloc[row_idx, self.status_col_index] = "不完全匹配"
    
    def _add_new_rows_to_template(self, new_companies: List[Dict], source_type: str):
        """在模板末尾添加新行"""
        if not new_companies:
            return
        
        # 创建新行数据
        new_rows = []
        for company in new_companies:
            new_row = [None] * len(self.template_df.columns)
            
            # 基座在线监测才填 D/E，其它源不填
            if source_type == '在线监测':
                new_row[3] = company['name']  # D列：企业名称
                new_row[4] = company['credit_code']  # E列：统一社会信用代码
            
            # 根据源类型设置对应列
            if source_type == '用电监控':
                new_row[7] = company['name']  # H列
                new_row[8] = company['credit_code']  # I列
            elif source_type == '排污许可':
                new_row[9] = company['name']  # J列
                new_row[10] = company['credit_code']  # K列
                new_row[11] = company.get('permit_number', '')  # L列
            elif source_type == '减排清单':
                new_row[12] = company['name']  # M列
                new_row[13] = company['credit_code']  # N列
            elif source_type == '源清单':
                new_row[14] = company['name']  # O列
                # 源清单没有信用代码，所以第15列留空
            
            new_rows.append(new_row)
        
        # 添加新行到模板
        new_df = pd.DataFrame(new_rows, columns=self.template_df.columns)
        self.template_df = pd.concat([self.template_df, new_df], ignore_index=True)
    
    def _fill_DE_from_temp_table(self):
        try:
            for item in self.temp_table:
                row_idx = item.get('template_row')
                if row_idx is not None and 0 <= row_idx < len(self.template_df):
                    self.template_df.iloc[row_idx, 3] = item.get('name', '')
                    self.template_df.iloc[row_idx, 4] = item.get('credit_code', '')
        except Exception as e:
            print(f"填充D/E列失败: {e}")
    
    def save_results(self, output_path: str):
        """保存结果：保留原模板表头与合并，仅覆盖第3行及之后的数据"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.template_path)
            ws = wb.active
            status_col_idx = (self.status_col_index + 1) if hasattr(self, 'status_col_index') else 17
            ws.cell(row=2, column=status_col_idx, value='匹配结果')
            current_max = ws.max_row
            if current_max >= 3:
                ws.delete_rows(3, current_max - 2)
            for r_idx, row_vals in enumerate(self.template_df.itertuples(index=False), start=3):
                for c_idx, val in enumerate(row_vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            wb.save(output_path)
            print(f"结果已保存到: {output_path}")
        except Exception as e:
            print(f"保存结果失败: {e}")
    
    def save_temp_table(self, csv_path: str):
        """保存临时表到CSV文件"""
        try:
            # 将临时表转换为DataFrame
            temp_df = pd.DataFrame(self.temp_table)
            
            # 保存到CSV文件
            temp_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"临时表已保存到: {csv_path}")
            
            # 打印临时表统计信息
            print(f"临时表统计:")
            print(f"  总企业数: {len(self.temp_table)}")
            
            # 按来源统计
            source_counts = {}
            for company in self.temp_table:
                source = company.get('source', '未知')
                source_counts[source] = source_counts.get(source, 0) + 1
            
            for source, count in source_counts.items():
                print(f"  {source}: {count}家")
                
        except Exception as e:
            print(f"保存临时表失败: {e}")
    
    def run_complete_matching(self, file_paths: Dict[str, str], output_path: str):
        """
        运行完整的匹配流程
        file_paths: {'用电监控': path, '排污许可': path, '减排清单': path, '源清单': path}
        """
        print("开始企业信息匹配流程...")
        self.start_time = datetime.datetime.now()
        
        # 按顺序匹配各个文件
        matching_order = ['用电监控', '排污许可', '减排清单', '源清单']
        
        for source_type in matching_order:
            if source_type in file_paths:
                self.temp_table = self.match_companies(file_paths[source_type], source_type)
            else:
                print(f"跳过{source_type}，文件路径未提供")
        
        # 根据临时表填充D/E列
        self._fill_DE_from_temp_table()
        # 根据临时表填充D/E列
        self._fill_DE_from_temp_table()
        # 根据临时表填充D/E列
        self._fill_DE_from_temp_table()
        # 根据临时表填充D/E列
        self._fill_DE_from_temp_table()
        # 根据临时表填充D/E列
        self._fill_DE_from_temp_table()
        # 根据临时表填充D/E列
        self._fill_DE_from_temp_table()
        # 根据临时表填充D/E列
        self._fill_DE_from_temp_table()
        # 保存最终结果
        self.save_results(output_path)
        
        # 保存临时表到CSV文件
        self.save_temp_table(output_path.replace('.xlsx', '_临时表.csv'))
        
        print("匹配流程完成！")
        print(f"临时表已保存到: {output_path.replace('.xlsx', '_临时表.csv')}")
        
        # 计算并显示总运行时间
        total_time = (datetime.datetime.now() - self.start_time).total_seconds()
        hours = int(total_time // 3600)
        minutes = int((total_time % 3600) // 60)
        seconds = total_time % 60
        
        print(f"\n程序总运行时间: {hours}小时 {minutes}分钟 {seconds:.2f}秒")
        print(f"总耗时: {total_time:.2f}秒")

def main():
    """主函数"""
    # 文件路径配置
    template_path = "/home/wjw/diting/ZhengHe/xingxuqiu/四川环科院结果模板2.xlsx"
    output_path = "/home/wjw/diting/ZhengHe/xingxuqiu/结果/四川环科院结果模板3_匹配结果1.xlsx"
    
    # 阿里云DashScope API密钥（请替换为您的实际密钥）
    dashscope_api_key = os.getenv('DASHSCOPE_API_KEY', 'your-dashscope-api-key-here')
    
    file_paths = {
        '用电监控': '/home/wjw/diting/ZhengHe/WenJianZhengHe/0 企业电监控档案导出.xlsx',
        '排污许可': '/home/wjw/diting/ZhengHe/WenJianZhengHe/0 排污许可证涉气企业.xlsx',
        '减排清单': '/home/wjw/diting/ZhengHe/xingxuqiu/减排清单_去重.xlsx',
        '源清单': '/home/wjw/diting/ZhengHe/WenJianZhengHe/0 源清单企业名单.xlsx'
    }
    
    # 初始化匹配系统
    matcher = EnterpriseMatchingSystem(template_path, dashscope_api_key)
    
    # 运行匹配流程
    matcher.run_complete_matching(file_paths, output_path)

if __name__ == "__main__":
    main()